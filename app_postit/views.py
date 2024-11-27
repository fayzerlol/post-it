from datetime import datetime, timedelta
import os
from django.shortcuts import render, redirect, get_object_or_404
import openpyxl
from .forms import FormularioUploadPlanilha
from .models import Planilha, SelecaoColunas, Quadro, Card
import pandas as pd
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseRedirect, JsonResponse
import io
from django.core.files.storage import FileSystemStorage
import urllib.parse
from django.conf import settings
from django.urls import reverse
from django.db import transaction

from .models import Planilha, Quadro
import json

def home(request):
    return render(request, 'home.html')

from django.core.files.storage import FileSystemStorage
import pandas as pd
from .models import Quadro, Planilha

from openpyxl import load_workbook

def upload_planilha(request):
    quadros = Quadro.objects.all()
    
    if request.method == 'POST':
        quadro_id = request.POST.get('quadro_id')
        if quadro_id:
            request.session['quadro_id'] = quadro_id  # Salva o ID do quadro na sessão

        if request.FILES.get('planilha'):
            planilha = request.FILES['planilha']
            fs = FileSystemStorage()
            filename = fs.save(planilha.name, planilha)
            uploaded_file_url = fs.url(filename)

            # Salva o caminho da planilha na sessão
            request.session['planilha_path'] = fs.path(filename)

            try:
                # Lendo os valores calculados da planilha
                wb = load_workbook(fs.path(filename), data_only=True)
                sheet = wb.active

                # Convertendo os dados para um DataFrame
                data = [[cell.value for cell in row] for row in sheet.iter_rows()]
                columns = data[0]  # Primeira linha como nomes das colunas
                rows = data[1:]    # Restante como dados
                df = pd.DataFrame(rows, columns=columns)

            except Exception as e:
                print(f"Erro ao processar o arquivo: {e}")
                return render(request, 'erro.html', {'mensagem': 'Erro ao carregar o arquivo. Verifique o formato ou o conteúdo.'})

            return render(request, 'selecionar_colunas.html', {
                'colunas': df.columns,
                'planilha_url': uploaded_file_url,
                'quadros': quadros,
                'quadro_id': quadro_id
            })

    return render(request, 'upload_planilha.html', {'quadros': quadros})



def selecionar_colunas(request):
    if request.method == 'POST':
        planilha_url = request.POST.get('planilha_url')

        if not planilha_url:
            return HttpResponse("URL da planilha não foi fornecida.", status=400)

        # Decodifica a URL para obter o nome do arquivo correto
        planilha_url_decoded = urllib.parse.unquote(planilha_url)
        
        # Constrói o caminho absoluto do arquivo
        planilha_path = os.path.join(settings.MEDIA_ROOT, os.path.basename(planilha_url_decoded))
        
        # Exibe o caminho completo do arquivo no terminal para depuração
        print(f"Caminho da planilha: {planilha_path}")
        
        # Verifica se o arquivo existe no caminho
        if not os.path.exists(planilha_path):
            return render(request, 'erro.html', {'mensagem': 'Arquivo não encontrado. Por favor, faça o upload novamente.'})

        try:
            # Carrega o arquivo Excel
            df = pd.read_excel(planilha_path)
            colunas = df.columns.tolist()
            return render(request, 'exibir_dados.html', {'colunas': colunas, 'dados': df.to_html(classes='table table-striped')})
        except Exception as e:
            print(f"Erro ao carregar a planilha: {e}")
            return render(request, 'erro.html', {'mensagem': 'Erro ao carregar a planilha.'})

    return render(request, 'upload.html')



def exibir_dados(request):
    planilha = Planilha.objects.latest('id')
    df = pd.read_excel(planilha.arquivo.path)

    colunas_selecionadas = request.POST.getlist('colunas')
    dados_filtrados = df[colunas_selecionadas]

    html_tabela = dados_filtrados.to_html(classes='table table-striped table-bordered')

    return render(request, 'exibir_dados.html', {'dados': html_tabela})


def exportar_dados(request):
    planilha = Planilha.objects.latest('id')
    df = pd.read_excel(planilha.arquivo.path)

    colunas_selecionadas = request.POST.getlist('colunas')
    dados_filtrados = df[colunas_selecionadas]

    output = io.StringIO()
    dados_filtrados.to_csv(output, index=False)
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=dados_filtrados.csv'

    return response


def listar_quadros(request):
    quadros = Quadro.objects.all()
    return render(request, 'listar_quadros.html', {'quadros': quadros})


def criar_quadros(request):
    colunas_selecionadas = SelecaoColunas.objects.values_list('colunas', flat=True).last().split(',')

    planilha_path = request.session.get('planilha_url')
    if not planilha_path or not os.path.exists(planilha_path):
        return render(request, 'erro.html', {'mensagem': 'Arquivo não encontrado. Por favor, faça o upload novamente.'})

    df = pd.read_excel(planilha_path, usecols=colunas_selecionadas)

    quadro_id = request.session.get('quadro_id')
    if not quadro_id:
        return render(request, 'erro.html', {'mensagem': 'Quadro não encontrado. Por favor, selecione um quadro.'})

    quadro = get_object_or_404(Quadro, id=quadro_id)
    print(f"Criando dados para o quadro: {quadro.nome}")

    for _, row in df.iterrows():
        dados_card = {coluna: str(row[coluna]) for coluna in colunas_selecionadas}
        conteudo_json = json.dumps(dados_card)  # Converte para JSON
        print(f"Criando card com dados: {conteudo_json}")
        Card.objects.create(quadro=quadro, titulo=row[colunas_selecionadas[0]], conteudo=conteudo_json)

    return render(request, 'exibir_quadro.html', {'quadros': Quadro.objects.all()})



from datetime import datetime, timedelta

def exibir_quadro(request, quadro_id):
    quadro = get_object_or_404(Quadro, id=quadro_id)
    cards = quadro.cards.all()
    colunas_importadas = quadro.colunas_importadas or []

    # DEBUG: Informações gerais do quadro
    print(f"Quadro ID: {quadro.id}, Nome: {quadro.nome}")
    print(f"Coluna de Status (definida): {quadro.coluna_status}")
    print(f"Coluna de Data (definida): {quadro.coluna_data}")

    # Normaliza o nome das colunas
    coluna_status_normalizada = (
        quadro.coluna_status.replace(" ", "").lower() if quadro.coluna_status else None
    )
    coluna_data_normalizada = (
        quadro.coluna_data.replace(" ", "").lower() if quadro.coluna_data else None
    )

    # Função para tentar converter a data com diferentes formatos
    def tentar_converter_data(data_str):
        formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]
        for formato in formatos:
            try:
                return datetime.strptime(data_str, formato)
            except ValueError:
                continue
        return None  # Retorna None se nenhum formato for compatível

    # Filtrar por status (apenas se a coluna de status estiver configurada)
    if coluna_status_normalizada:
        filtered_cards = []
        for card in cards:
            conteudo_normalizado = {key.replace(" ", "").lower(): value for key, value in card.conteudo.items()}
            if coluna_status_normalizada in conteudo_normalizado:
                status_value = str(conteudo_normalizado[coluna_status_normalizada]).strip().lower()
                if status_value in ["aberto", "iniciado"]:
                    filtered_cards.append(card)
        cards = filtered_cards

    # Filtro por intervalo de datas
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    if data_inicio and data_fim and coluna_data_normalizada:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, "%d-%m-%Y")
            data_fim_dt = datetime.strptime(data_fim, "%d-%m-%Y")
            filtered_cards = []
            for card in cards:
                conteudo_normalizado = {key.replace(" ", "").lower(): value for key, value in card.conteudo.items()}
                if coluna_data_normalizada in conteudo_normalizado:
                    card_data = tentar_converter_data(conteudo_normalizado[coluna_data_normalizada])
                    if card_data and data_inicio_dt <= card_data <= data_fim_dt:
                        filtered_cards.append(card)
            cards = filtered_cards
        except ValueError as e:
            print(f"Erro ao processar o filtro de datas: {e}")  # DEBUG

    # Aplicar lógica de cores e status
    today = datetime.now()
    for card in cards:
        card.cor = 'default'
        card.status_data = 'no_prazo'
        conteudo_normalizado = {key.replace(" ", "").lower(): value for key, value in card.conteudo.items()}

        # Verifica o status "Iniciado" primeiro
        if coluna_status_normalizada and coluna_status_normalizada in conteudo_normalizado:
            status_value = str(conteudo_normalizado[coluna_status_normalizada]).strip().lower()
            if status_value == "iniciado":
                card.cor = 'verde'
                card.status_data = 'iniciado'

        # Aplica lógica de data apenas se o status não for "iniciado"
        if card.status_data != 'iniciado' and coluna_data_normalizada:
            if coluna_data_normalizada in conteudo_normalizado:
                card_data_str = conteudo_normalizado[coluna_data_normalizada]
                card_data = tentar_converter_data(card_data_str)
                if card_data:
                    card.conteudo[quadro.coluna_data] = card_data.strftime("%d-%m-%Y")  # Formata para exibição
                    if card_data < today:
                        card.cor = 'vermelho'
                        card.status_data = 'vencido'
                    elif today <= card_data <= today + timedelta(days=10):
                        card.cor = 'amarelo'
                        card.status_data = 'quase_vencido'
                    elif card_data > today + timedelta(days=10):
                        card.cor = 'default'
                        card.status_data = 'no_prazo'

    # Ordenação dos cards
    cards = sorted(cards, key=lambda x: (
        0 if x.status_data == 'iniciado' else  # Prioridade para "iniciado"
        1 if x.status_data == 'vencido' else  # Depois os vencidos
        2 if x.status_data == 'quase_vencido' else  # Depois os quase vencidos
        3,  # Por fim, os cards sem status
        tentar_converter_data(x.conteudo.get(quadro.coluna_data, "")) or datetime.max  # Ordena pela data
    ))

    # Passa o contexto para o template
    context = {'quadro': quadro, 'cards': cards, 'colunas_importadas': colunas_importadas}
    return render(request, 'exibir_quadro.html', context)






def importar_planilha(planilha_path, quadro_id, coluna_identificadora, colunas_importadas, colunas_no_card):
    # Carrega a planilha
    workbook = openpyxl.load_workbook(planilha_path)
    sheet = workbook.active

    # Normaliza o cabeçalho para remover caracteres invisíveis, como \r, \n e espaços extras
    cabecalho = [str(cell.value).strip().replace('\r', '').replace('\n', '').replace(' ', '') for cell in sheet[1]]
    
    # Aplica a mesma normalização nas colunas que foram passadas para importar
    colunas_importadas = [col.strip().replace('\r', '').replace('\n', '').replace(' ', '') for col in colunas_importadas]
    colunas_no_card = [col.strip().replace('\r', '').replace('\n', '').replace(' ', '') for col in colunas_no_card]
    coluna_identificadora = coluna_identificadora.strip().replace('\r', '').replace('\n', '').replace(' ', '')

    print(f"Cabecalho lido: {cabecalho}")
    print(f"Colunas Importadas após limpeza: {colunas_importadas}")

    # Mapeia os índices das colunas importadas com base no cabeçalho normalizado
    try:
        coluna_indices = {coluna: cabecalho.index(coluna) for coluna in colunas_importadas}
    except ValueError as e:
        raise ValueError(f"Erro ao procurar colunas: {e}. Verifique se os nomes das colunas na planilha correspondem exatamente aos nomes fornecidos.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Cria um dicionário com as colunas importadas
        dados = {}
        for coluna in colunas_importadas:
            valor = row[coluna_indices[coluna]]
            if isinstance(valor, datetime):
                valor = valor.strftime('%Y-%m-%d')  # Converte datetime para string
            dados[coluna] = valor
        
        # Identifica a chave única para o título
        titulo = dados.get(coluna_identificadora)
        
        if not titulo:
            raise ValueError(f"O valor da coluna identificadora '{coluna_identificadora}' é nulo ou não encontrado para a linha {row}. Verifique se a planilha está correta.")

        # Verifica se já existe um item com esse identificador
        item_existente = Card.objects.filter(titulo=titulo, quadro_id=quadro_id).first()

        if item_existente:
            # Se existir, atualiza os dados
            item_existente.conteudo = dados
            item_existente.save()
        else:
            # Se não existir, cria um novo
            novo_item = Card(titulo=titulo, conteudo=dados, quadro_id=quadro_id)
            novo_item.save()

    return True


def salvar_colunas(request):
    if request.method == 'POST':
        quadro_id = request.POST.get('quadro_id')
        planilha_url = request.POST.get('planilha_url')
        
        if not planilha_url:
            return HttpResponse("URL da planilha não foi fornecida.", status=400)
        
        # Decodifica a URL e monta o caminho completo
        planilha_url_decoded = urllib.parse.unquote(planilha_url.lstrip('/media/'))
        planilha_path = os.path.join(settings.MEDIA_ROOT, os.path.basename(planilha_url_decoded))
        
        # Verifica se o arquivo realmente existe
        if not os.path.exists(planilha_path):
            print(f"Erro: O arquivo {planilha_path} não foi encontrado.")
            return HttpResponse(f"Erro ao processar a planilha: Arquivo {planilha_path} não encontrado.", status=404)

        coluna_identificadora = request.POST.get('coluna_identificadora')
        colunas_importadas = request.POST.getlist('colunas_importadas')
        colunas_no_card = request.POST.getlist('colunas_no_card')
        
        # Log para verificar as colunas recebidas
        print(f"Coluna Identificadora: {coluna_identificadora}")
        print(f"Colunas Importadas: {colunas_importadas}")
        print(f"Colunas no Card: {colunas_no_card}")
        
        try:
            # Recupera o quadro pelo ID
            quadro = get_object_or_404(Quadro, id=quadro_id)
            
            # Salva as colunas importadas e as colunas selecionadas para exibição no card
            quadro.colunas_importadas = colunas_importadas  # Se você quiser salvar essas também
            quadro.colunas_no_card = colunas_no_card
            
            # Salva as alterações no banco de dados
            quadro.save()
            
            # Processa a planilha usando a função importar_planilha
            importar_planilha(planilha_path, quadro_id, coluna_identificadora, colunas_importadas, colunas_no_card)
        except Exception as e:
            print(f"Ocorreu um erro ao processar a planilha: {str(e)}")
            return HttpResponse(f"Ocorreu um erro ao processar a planilha: {str(e)}", status=500)
        
        return redirect('listar_quadros')
    else:
        return HttpResponse("Método não permitido.", status=405)



def listar_cards(request):
    cards = Card.objects.all()
    return render(request, 'listar_cads.html', {'cards': cards})

def pagina_inicial(request):
    quadros = Quadro.objects.all()
    return render(request, 'pagina_inicial.html', {'quadros':quadros})

def criar_quadro(request):
    if request.method == 'POST':
        nome_quadro = request.POST.get('nome_quadro')
        
        if nome_quadro:
            # Cria o quadro com valores padrão para colunas_importadas e colunas_no_card
            Quadro.objects.create(nome=nome_quadro, colunas_importadas="", colunas_no_card="")
            return redirect('upload_planilha')
    
    return render(request, 'criar_quadro.html')


def editar_quadro(request, quadro_id):
    quadro = get_object_or_404(Quadro, id=quadro_id)
    colunas_importadas = quadro.colunas_importadas  # Certifique-se que existe este campo ou método

    if request.method == 'POST':
        # Atualizar os dados do quadro com as entradas do formulário
        quadro.nome = request.POST.get('nome_quadro')
        quadro.coluna_status = request.POST.get('coluna_status')
        quadro.coluna_data = request.POST.get('coluna_data')
        quadro.save()  # Salvar as alterações no banco de dados
        
        # Redirecionar ou exibir uma mensagem de sucesso
        return redirect('listar_quadros')  # Altere para a URL de listagem de quadros ou outra apropriada

    context = {
        'quadro': quadro,
        'colunas_disponiveis': colunas_importadas,  # Passar as colunas para o template
    }
    return render(request, 'editar_quadro.html', context)

def deletar_quadro(request, quadro_id):
    quadro = get_object_or_404(Quadro, id=quadro_id)
    if request.method == 'POST':
        quadro.delete()
        return redirect('listar_quadros')
    return render(request, 'deletar_quadro.html', {'quadro': quadro})


def salvar_personalizado(request, quadro_id):
    if request.method == 'POST':
        quadro = get_object_or_404(Quadro, id=quadro_id)
        quadro.tamanho_fonte = request.POST.get('tamanho_fonte', quadro.tamanho_fonte)
        quadro.cor_background = request.POST.get('cor_background', quadro.cor_background)
        quadro.cor_fonte = request.POST.get('cor_fonte', quadro.cor_fonte)
        quadro.cor_cards = request.POST.get('cor_cards', quadro.cor_cards)
        quadro.borda_cards = request.POST.get('borda_cards', quadro.borda_cards)
        quadro.sombra_cards = request.POST.get('sombra_cards', quadro.sombra_cards)
        quadro.save()
        
        return redirect('exibir_quadros', quadro_id=quadro.id)
    
    else:
        return HttpResponse("Método não permitido.", status=405)


@transaction.atomic
def atualizar_planilha(request, quadro_id):
    quadro = get_object_or_404(Quadro, id=quadro_id)

    if request.method == "POST":
        # Obtém a planilha do formulário
        planilha = request.FILES.get("planilha")
        if not planilha:
            return HttpResponseBadRequest("Nenhuma planilha foi enviada.")

        try:
            # Carrega a planilha usando pandas
            df = pd.read_excel(planilha)

            # Verifica se as colunas importadas ainda existem na nova planilha
            colunas_importadas = quadro.colunas_importadas or []
            colunas_no_card = quadro.colunas_no_card or []

            if not colunas_importadas:
                return HttpResponseBadRequest("Nenhuma coluna importada foi definida para este quadro.")

            # Define a primeira coluna como identificadora
            coluna_identificadora = colunas_importadas[0]

            # Verifica se a coluna identificadora e as colunas importadas existem na nova planilha
            colunas_faltantes = [col for col in colunas_importadas if col not in df.columns]
            if colunas_faltantes:
                return HttpResponseBadRequest(
                    f"As seguintes colunas estão faltando na nova planilha: {', '.join(colunas_faltantes)}"
                )

            # Normaliza os dados da planilha (converte valores para strings e formata datas)
            def normalizar_valor(valor):
                if pd.isna(valor):  # Ignora valores NaN
                    return None
                elif isinstance(valor, (pd.Timestamp, datetime)):  # Converte Timestamp ou datetime para string
                    return valor.strftime("%Y-%m-%d")
                return str(valor).strip()

            # Atualiza os dados do quadro
            for _, row in df.iterrows():
                # Identificador único (chave primária para identificar o item)
                identificador = normalizar_valor(row.get(coluna_identificadora))
                if not identificador:
                    continue  # Pula itens sem identificador

                # Busca um card existente
                card = Card.objects.filter(quadro=quadro, titulo=identificador).first()

                # Converte os dados da linha para um dicionário com apenas as colunas importadas
                dados_card = {coluna: normalizar_valor(row[coluna]) for coluna in colunas_importadas}

                if card:
                    # Atualiza o card existente
                    card.conteudo = dados_card
                    card.save()
                else:
                    # Cria um novo card
                    Card.objects.create(
                        quadro=quadro,
                        titulo=identificador,
                        conteudo=dados_card
                    )

            return redirect("exibir_quadros", quadro_id=quadro.id)

        except Exception as e:
            print(f"Erro ao atualizar a planilha: {e}")
            return HttpResponseBadRequest("Ocorreu um erro ao processar a planilha. Verifique o formato.")

    return HttpResponseBadRequest("Método não permitido.")

